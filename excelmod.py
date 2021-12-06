from settings import Config
from s2t_class import Speech2Text, test
import openpyxl
import json
import sys
import traceback
import os


class ModExcel:
    def __init__(self, path="samples/samples.xlsx",
                 config=Config().get_config()):
        self.path = path
        self.workbook = openpyxl.load_workbook(path)
        self.row = 1
        self.col = 1
        self.config = config
        self.sheet = self.workbook.active

    def getSheet(self):
        return self.sheet

    def getCell(self):
        try:
            return self.sheet.cell(row=self.row, column=self.col)
        except Exception:
            return None

    def locate(self, *args):
        flag = 0
        for i in range(1, self.sheet.max_row + 1):
            for j in range(1, self.sheet.max_column + 1):
                for k in args:
                    cv = self.sheet.cell(row=i, column=j)
                    if k in cv.value:
                        print(
                            f"\n\n{k} found in {cv.address} having {cv.value}",
                            end="\n\n"
                        )
                        flag += 1

        if flag == 0:
            print("\n\nSorry, no result found.\n\n")

    def getDetails(self, *args):
        if args[0] == "address":
            print(
                f"\n\nCurrent cell address is {self.getCell().coordinate}\n\n"
            )
            return self.getCell().coordinate
        elif args[0] == "value":
            print(f"\n\nCurrent cell value is {self.getCell().value}\n\n")
            return self.getCell().value
        elif args[0] == "sheet":
            print(f"\n\n\n\nCurrent sheet is {self.getSheet()}\n\n\n\n")
        elif args[0] == "workbook":
            print(f"\n\n\n\nCurrent workbook is {self.getSheet()}\n\n\n\n")

    def navigate(self, args):
        # try:
        for i in args:
            if i.lower() == "right":
                print("\n\nMoving one column right\n\n")
                self.col += 1
            elif i.lower() == "left":
                print("\n\nMoving one column left\n\n")
                self.col -= 1
            elif i.lower() == "up":
                print("\n\nMoving one row up\n\n")
                self.row -= 1
            elif i.lower() == "down":
                print("\n\nMoving one row down\n\n")
                self.row += 1
            else:
                print("Sorry, command not recognized :(")

        # self.getCell()
        # except:
        #     return None

    def setValue(self, cmd):
        val = cmd[0]
        cell = self.getCell()
        try:
            cell.value = val
            print("\n\nModified the concerned excel file.\n\n")
            self.workbook.save(self.path)
            return True
        except Exception:
            print("\n\nCouldn't modify the concerned excel file.\n\n")
            return False


def parsecmd(cmd):
    # Enter ABCD and move left and down and get cell and value
    # import json

    # a = input("\n\nEnter a string to generate the tree of:\n\n")

    tree = [[y.strip() for y in x.strip().split()] for x in cmd.split("and ")]
    tree2 = []

    for i in range(len(tree)):
        if len(tree[i]) == 1:
            tree2[-1] += list(tree[i])
        else:
            tree2.append(tree[i])

        # print(tree2)

    for i in tree2:
        for j in range(len(i)):
            if j == 0:
                i[j] = i[j].lower()

    print(f"\n\nCrudely Parsed Tree: {tree2}")

    tree3 = []

    keyword = None
    arg = None
    tpl = None

    for i in tree2:
        if i[0] == "enter":
            keyword = i[0]
            arg = " and ".join(i[1:])
            tpl = tuple((keyword, arg))

        else:  # if i[0] == 'move':
            keyword = i[0]
            arg = i[1:]
            tpl = []
            for i in range(len(arg)):
                tpl += [tuple((keyword, arg[i]))]

        if type(tpl) == type(tuple()):
            tree3.append(tpl)
        elif type(tpl) == type([]):
            tree3 += tpl
        else:
            continue

    print(f"\n\nRefined Parsed Tree: {json.dumps(tree3, indent=4)}")
    del tree
    del tree2

    return tree3


def execute(xl, cfg, parsed_tree):
    # parsed_tree = parsecmd(text)
    commands = list([x[0] for x in parsed_tree])
    attributes = list([x[1] for x in parsed_tree])

    print(f"\n\nParsed syntax: {parsed_tree}\n\n")
    print(f"\n\nParsed commands: {commands}\n\n")

    print("\n\nConfiguration available:\n\n")
    for i in cfg.keys():
        print(f"{i}\t{cfg[i]}")

    for i in range(len(commands)):
        if commands[i] in cfg.keys():
            print(f"\n\nCommand recognized: {commands[i]}.")
            print(f"Parameters: {attributes[i]}\n\n")
            # if commands[i] not in ["get"]:
            cmd = f"xl.{cfg[commands[i]]}(['{attributes[i]}'])"

            # else:
            #     cmd = f"xl.{cfg[commands[i]]}()"

            print(f"\n\nBuilt command: {cmd}\n\n")
            exec(cmd)
        else:
            print(
                f"\n\nCommand {commands[i]} isn't valid... \n\n"
            )


def driver_excelmod():
    s2t = Speech2Text(mode="mic", src="vosk", output="a_test.txt")
    # s2t = Speech2Text(mode="rec", src="vosk", output="a_test.txt")
    cfg = Config().get_config()
    xl = ModExcel(config=cfg)
    cmd = None

    while True:
        try:
            if s2t.src == "vosk":
                text = json.loads(s2t.listen())["text"]
            elif s2t.src == "google":
                text = s2t.listen()
            print(f"\n\nCommand given: {text}\n\n")

            execute(xl, cfg, parsecmd(text))
            # parsed_tree = parsecmd(text)
            # commands = list([x[0] for x in parsed_tree])
            # attributes = list([x[1] for x in parsed_tree])

            # print(f"\n\nParsed syntax: {parsed_tree}\n\n")
            # print(f"\n\nParsed commands: {commands}\n\n")

            # print("\n\nConfiguration available:\n\n")
            # for i in cfg.keys():
            #     print(f"{i}\t{cfg[i]}")

            # Execution block
            # execute(commands, attributes, cfg)
            # for i in range(len(commands)):
            #     if commands[i] in cfg.keys():
            #         print(f"\n\nCommand recognized: {commands[i]}.")
            #         print(f"Parameters: {attributes[i]}\n\n")
            #         # if commands[i] not in ["get"]:
            #         cmd = f"xl.{cfg[commands[i]]}(['{attributes[i]}'])"

            #         # else:
            #         #     cmd = f"xl.{cfg[commands[i]]}()"

            #         print(f"\n\nBuilt command: {cmd}\n\n")
            #         exec(cmd)
            #     else:
            #         print(
            #             f"\n\nCommand {commands[i]} isn't valid... \n\n"
            #         )
        except Exception as e:
            traceback.print_exc()
            print("\n\nAn error occured. Error details: \t", e)


def test_excelmod():
    test_inst_vosk = Speech2Text(
        mode="rec",
        loc="/media/anuran/Samsung SSD 970 EVO 1TB/Projects/NLP Project/tests/input/Instructions-for-the-voice-recog-system.wav",
        src="vosk",
        output="/media/anuran/Samsung SSD 970 EVO 1TB/Projects/NLP Project/tests/output/random_test.txt"
    )

    text = json.loads(test_inst_vosk.listen())["text"]

    cfg = Config().get_config()
    xl = ModExcel(config=cfg)

    execute(xl, cfg, parsecmd(text))
    os.remove("/media/anuran/Samsung SSD 970 EVO 1TB/Projects/NLP Project/tests/output/random_test.txt")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        # test()
        test_excelmod()
    else:
        driver_excelmod()
